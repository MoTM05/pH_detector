import os
import numpy as np
import cv2
import matplotlib.pyplot as plt
import openpyxl
import tkinter as tk
from PIL import Image, ImageDraw, ImageFont, ImageTk
from tkinter import Tk, filedialog, messagebox
from datetime import datetime
from pillow_heif import register_heif_opener


# Функция для создания градиента (предыдущий код)

def create_gradient(start_color, end_color, width, height):
    gradient = np.zeros((height, width, 3), dtype=np.uint8)
    for i in range(width):
        alpha = i / width
        color = (1 - alpha) * np.array(start_color) + alpha * np.array(end_color)
        gradient[:, i] = color
    return gradient


# Функция для нахождения наиболее ближайшего цвета
def find_closest_color(gradient, target_rgb):
    # Конвертация RGB в BGR
    target_bgr = target_rgb[::-1]

    # Расчет Евклидова расстояния от каждого пикселя до целевого цвета
    distances = np.sqrt(np.sum((gradient - target_bgr) ** 2, axis=2))

    # Находим индекс пикселя с минимальным расстоянием
    min_dist_index = np.unravel_index(np.argmin(distances), distances.shape)

    # Возвращаем цвет пикселя и его позицию
    closest_color = gradient[min_dist_index]
    return closest_color[::-1], min_dist_index  # Возвращаем в формате RGB и позицию


def create_gradient_segment(start_color, end_color, width):
    """Создает горизонтальный градиент между двумя заданными цветами."""
    start_color = np.array(start_color[::-1], dtype=np.uint8)  # Конвертация из RGB в BGR
    end_color = np.array(end_color[::-1], dtype=np.uint8)  # Конвертация из RGB в BGR

    gradient = np.zeros((1, width, 3), dtype=np.uint8)

    for x in range(width):
        alpha = x / (width - 1)
        color = (1 - alpha) * start_color + alpha * end_color
        gradient[0, x] = color

    return gradient


def create_multi_gradient(colors, width, height):
    """Создает градиент через список цветов."""
    segments_count = len(colors) - 1
    segment_width = width // segments_count
    gradient = np.zeros((height, width, 3), dtype=np.uint8)

    for i in range(segments_count):
        segment = create_gradient_segment(colors[i], colors[i + 1], segment_width)
        segment = cv2.resize(segment, (segment_width, height), interpolation=cv2.INTER_LINEAR)
        gradient[:, i * segment_width:(i + 1) * segment_width] = segment

    # Для последнего сегмента, если ширина не делится нацело
    if width % segments_count != 0:
        last_segment_width = width - (segments_count - 1) * segment_width
        last_segment = create_gradient_segment(colors[-2], colors[-1], last_segment_width)
        last_segment = cv2.resize(last_segment, (last_segment_width, height), interpolation=cv2.INTER_LINEAR)
        gradient[:, -last_segment_width:] = last_segment

    return gradient


def process_images_from_folder(folder_path):
    results = []  # Список для хранения результатов

    # Перебор всех файлов в директории
    for filename in os.listdir(folder_path):
        if filename.endswith(('.png', '.jpg', '.jpeg')):  # Проверка формата файла
            file_path = os.path.join(folder_path, filename)
            image = Image.open(file_path)
            width, height = image.size

            # Вычисление координат для квадрата 50x50 пикселей в центре изображения
            left = (width / 2) - 50
            top = (height / 2) - 50
            right = (width / 2) + 50
            bottom = (height / 2) + 50

            # Обрезка изображения
            cropped_image = image.crop((int(left), int(top), int(right), int(bottom)))
            # Вычисление средних значений RGB
            average_color = [c for c in cropped_image.resize((1, 1), Image.Resampling.LANCZOS).getpixel((0, 0))]

            # Добавление результатов в список
            results.append([filename.rsplit('.', 1)[0], average_color[0], average_color[1], average_color[2]])

    return results


def add_ph_value_as_text(image, res):
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    text_color = (255, 0, 0)
    image_width, image_height = image.size
    text_position = (image_width - 50, 10)

    draw.text(text_position, f"pH: {res}", font=font, fill=text_color)

    return image


def open_file_dialog():
    root = Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.heic;")])

    if file_path:
        return file_path
    else:
        return None


# получаем ID эксперимента в таблице эксель для корректной записи
def get_experiment_id(excel_file_path):
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        return sheet.max_row
    except FileNotFoundError:
        return 1


def get_image_name(image_path):
    return os.path.basename(image_path)


def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x_coordinate = (screen_width - width) // 2
    y_coordinate = (screen_height - height) // 2
    window.geometry(f"{width}x{height}+{x_coordinate}+{y_coordinate}")


        # Перезапускаем программу


def process_image(root):
    print("Идет обработка изображения. Подождите...")
    image_path = open_file_dialog()
    if not image_path:
        print("Отменено пользователем.")
        return

    filename, ext = os.path.splitext(image_path)
    ext = ext.lower()

    if ext == '.heic' or ext == '.HEIC':  # проверка на .heic
        register_heif_opener()  # какая-то функция для считывания изображений формата .heic
        image = Image.open(image_path)  # открываем эту фотку при помощи PIL
        image.save(filename + '.png')  # пересохраняем фотку в формат .png
        img = cv2.imread(filename + '.png')  # открываем фотку уже в .png формате и работаем с ней
    else:
        img = cv2.imread(image_path)

    if img is None:
        print("Ошибка: Не удалось загрузить изображение. Проверьте путь к файлу.")
        return

    for widget in root.winfo_children():
        widget.destroy()

    gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blurred_img = cv2.GaussianBlur(gray_img, (9, 9), 2)
    circles = cv2.HoughCircles(blurred_img, cv2.HOUGH_GRADIENT, dp=3, minDist=20, param1=50, param2=30)

    if circles is not None:
        circles = circles[0]  # Преобразование координат кругов в целочисленный формат
        # Перебираем обнаруженные круги
        for circle in circles:
            x, y, r = circle.astype(int)
            # Вырезаем круг из исходного изображения
            cropped_circle = img[y - r:y + r, x - r:x + r]
            # Сохраняем вырезанный круг
            cv2.imwrite("cropped_circle.png", cropped_circle)
            break

    else:
        print("Изображение некорректно. Пожалуйста, выберите другое изображение.")


        # открываем и показываем сохраненный вырезанный круг для дальнейшей работы и выреза квадрата
    img = Image.open('cropped_circle.png')
    img.show(title='Cropped circle')

    bar_width, bar_height = 800, 100

    gradient = create_multi_gradient(colors, bar_width, bar_height)
    width, height = img.size
    left = (width / 2) - 50
    top = (height / 2) - 50
    right = (width / 2) + 50
    bottom = (height / 2) + 50

    cropped_image = img.crop((int(left), int(top), int(right), int(bottom)))
    # вывод квадратного изображения
    cropped_image.show(title='Square image')

    average_color = [c for c in cropped_image.resize((1, 1), Image.Resampling.LANCZOS).getpixel((0, 0))]
    target_rgb = (average_color[0], average_color[1], average_color[2])
    closest_color, position = find_closest_color(gradient, target_rgb)
    r, g, b = closest_color[0], closest_color[1], closest_color[2]

    res = low_num + delta_ph * (position[1] / bar_width)
    res = round(res, 2)

    experiment_id = get_experiment_id('experiments.xlsx')
    image_name = get_image_name(image_path)
    pH_value = res
    current_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    try:
        workbook = openpyxl.load_workbook('experiments.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet.max_row == 1:
        sheet["A1"] = "ID"
        sheet["B1"] = "Шкала"
        sheet["C1"] = "Изображение"
        sheet["D1"] = "Red"
        sheet["E1"] = "Green"
        sheet["F1"] = "Blue"
        sheet["G1"] = "pH"
        sheet["H1"] = "Дата и время"

    row_number = sheet.max_row + 1

    sheet[f"A{row_number}"] = experiment_id
    sheet[f"B{row_number}"] = name
    sheet[f"C{row_number}"] = image_name
    sheet[f"D{row_number}"] = r
    sheet[f"E{row_number}"] = g
    sheet[f"F{row_number}"] = b
    sheet[f"G{row_number}"] = pH_value
    sheet[f"H{row_number}"] = current_datetime

    workbook.save('experiments.xlsx')

    cv2.line(gradient, (position[1], 0), (position[1], gradient.shape[0]), (0, 0, 255), thickness=2)
    font = cv2.FONT_HERSHEY_SIMPLEX
    res_text = f"pH: {res}"
    text_size = cv2.getTextSize(res_text, font, 1, 2)[0]
    text_x = (gradient.shape[1] - text_size[0]) // 2
    text_y = (gradient.shape[0] + text_size[1]) // 2
    cv2.putText(gradient, res_text, (text_x, text_y), font, 0.8, (0, 0, 255), 1, cv2.LINE_AA)

    gradient_bgr = cv2.cvtColor(gradient, cv2.COLOR_RGB2BGR)
    plt.figure(figsize=(15, 7))
    plt.imshow(gradient_bgr, interpolation='nearest')
    frame1 = plt.gca()
    frame1.axes.get_xaxis().set_visible(False)
    frame1.axes.get_yaxis().set_visible(False)
    plt.show()

    folder_path_2 = "results"
    if not os.path.exists(folder_path_2):
        os.makedirs(folder_path_2)

    time = datetime.now().strftime('%d %m %Y %H %M %S')
    cv2.imwrite(f'results/{name_en}_pH_{res}_{time}.png', gradient)

    os.remove('cropped_circle.png')
    if ext == '.heic' or ext == '.HEIC':
        os.remove(filename + '.png')

    messagebox.showinfo("Готово", "Обработка изображения завершена.")
    show_end_dialog(root)


def main():
    global colors

    def choose_bromothymol_blue():
        global colors, delta_ph, low_num, name, name_en
        colors = [(205, 188, 109),
                  (202, 188, 126),
                  (187, 177, 130),
                  (182, 174, 130),
                  (191, 188, 150),
                  (180, 180, 154),
                  (188, 193, 168),
                  (175, 186, 174),
                  (167, 182, 176),
                  (158, 178, 179),
                  (148, 172, 178),
                  (143, 170, 184),
                  (125, 162, 180)]
        delta_ph = 8 - 5.8
        low_num = 5.8
        name = 'Бромтимоловый синий'
        name_en = 'Bromothymol Blue'
        root.title("pH_detector")
        root.iconbitmap('photos/icon.ico')
        return colors

    def choose_bromocresol_purple():
        global colors, delta_ph, low_num, name, name_en
        colors = [(201, 193, 159),
                  (201, 195, 159),
                  (206, 201, 172),
                  (209, 204, 172),
                  (201, 198, 173),
                  (203, 199, 171),
                  (201, 197, 171),
                  (210, 204, 174),
                  (198, 196, 176),
                  (206, 200, 176),
                  (201, 197, 176),
                  (197, 191, 172),
                  (190, 184, 168),
                  (187, 184, 174),
                  (184, 181, 174),
                  (181, 181, 181),
                  (177, 176, 178),
                  (177, 177, 181),
                  (179, 178, 185),
                  (180, 178, 185),
                  (175, 176, 191),
                  (165, 166, 182),
                  (166, 167, 186),
                  (161, 161, 184),
                  (155, 156, 177)]
        delta_ph = 7.2 - 4.8
        low_num = 4.8
        name = 'Бромкрезоловый пурпурный'
        name_en = 'Bromocresol purple'
        root.title("pH_detector")
        root.iconbitmap('photos/icon.ico')

    root = tk.Tk()
    root.title("pH_detector")
    root.iconbitmap('photos/icon.ico')

    use_color_scale = tk.IntVar(value=1)
    scale_label = tk.Label(root, text="Выберите шкалу градиента:")
    scale_label.pack()

    color_scale_frame = tk.Frame(root)
    color_scale_frame.pack()

    def process_bromothymol():
        choose_bromothymol_blue()
        process_image(root)

    def process_bromocresol():
        choose_bromocresol_purple()
        process_image(root)

    bromothymol_button = tk.Button(color_scale_frame, text="Бромтимоловый синий", command=process_bromothymol)
    bromothymol_button.pack(side=tk.LEFT)

    bromocresol_button = tk.Button(color_scale_frame, text="Бромкрезоловый пурпурный", command=process_bromocresol)
    bromocresol_button.pack(side=tk.RIGHT)

    window_width = 400
    window_height = 100
    center_window(root, window_width, window_height)

    root.protocol("WM_DELETE_WINDOW", lambda: show_end_dialog(root))  # Обработка закрытия окна

    root.mainloop()

#start_button = tk.Button(root, text="Выбрать фотографию", command=lambda: process_image(root))
def show_end_dialog(root):
    result = messagebox.askquestion("Завершение программы", "Закрыть приложение?")
    if result == "yes":
        root.quit()
        root.destroy()
        return False
    else:
        root.destroy()  # Уничтожаем окно
        root.quit()     # Завершаем основной цикл Tkinter
        main()


if __name__ == "__main__":
    main()

